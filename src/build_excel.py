"""
build_excel.py
===============
Builds the Excel deliverable: a multi-tab workbook containing the
3-statement model, DCF, comps, and a summary "Dashboard" tab.

Color coding (industry standard, also documented in README):
    Blue  = hardcoded inputs / assumptions
    Black = formulas
    Green = links pulling from another sheet / source data in same workbook
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
import data_fetch
import three_statement_model as tsm
import valuation as val

BLUE = Font(color="0000FF")
BLACK = Font(color="000000")
GREEN = Font(color="008000")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
SUBTITLE_FONT = Font(italic=True, size=10, color="595959")
SECTION_FONT = Font(bold=True, size=12, color="1F4E78")
BOLD = Font(bold=True)

NUM_FMT = '#,##0;(#,##0);"-"'
AUD_FMT = '"A$"#,##0.00'
PCT_FMT = '0.0%;(0.0%);"-"'
MULT_FMT = '0.0"x"'


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_title(ws, title, subtitle, last_col=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = SUBTITLE_FONT


def write_table(ws, start_row, df, fmt=NUM_FMT, source_text=None):
    headers = [df.index.name or ""] + list(df.columns)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center" if j > 1 else "left")

    for i, (idx, row) in enumerate(df.iterrows(), start=start_row + 1):
        rc = ws.cell(row=i, column=1, value=idx)
        rc.font = BOLD
        for j, value in enumerate(row, start=2):
            cell = ws.cell(row=i, column=j, value=float(value))
            cell.number_format = fmt
            cell.font = BLACK

    end_row = start_row + len(df)
    if source_text:
        sc = ws.cell(row=end_row + 1, column=1, value=source_text)
        sc.font = Font(italic=True, size=8, color="808080")
    return end_row + 2


def derive_recommendation(dcf, comps):
    dcf_upside = dcf["upside_downside_pct"]
    comps_avg_implied = (comps["implied_price_ev_avg_aud"] + comps["implied_price_pe_avg_aud"]) / 2
    comps_upside = (comps_avg_implied / comps["current_share_price_aud"]) - 1
    blended = (dcf_upside + comps_upside) / 2

    if blended > 0.10:
        rec = "BUY"
    elif blended < -0.10:
        rec = "SELL"
    else:
        rec = "HOLD"

    rationale = (
        f"DCF implies {dcf_upside:.1%} vs current price; comps (avg of EV/EBITDA & P/E) imply {comps_upside:.1%}. "
        f"Blended view: {blended:.1%}. BHP trades at a premium to peers ({comps['bhp_current_ev_ebitda']:.1f}x "
        f"EV/EBITDA vs peer avg {comps['peer_avg_ev_ebitda']:.1f}x), reflecting its scale, low-cost iron ore "
        f"position and copper growth optionality - but this premium leaves limited margin of safety at current levels."
    )
    return rec, rationale, blended


def build_dashboard(wb, snapshot, model, dcf, comps):
    ws = wb.create_sheet("Dashboard", 0)
    set_col_widths(ws, [32, 16, 16, 16, 16, 16, 16, 16])
    company = snapshot["company"]

    write_title(ws, f"{company['name']} ({company['ticker']}) - Equity Research Dashboard",
                "Simplified equity research model | All financials in USD millions unless stated | Educational/portfolio purposes only - not investment advice")

    row = 4
    ws.cell(row=row, column=1, value="COMPANY SNAPSHOT").font = SECTION_FONT
    row += 1
    snapshot_rows = [
        ("Sector", company["sector"], None),
        ("Industry", company["industry"], None),
        ("Reporting Currency", company["reporting_currency"], None),
        ("Shares Outstanding (m)", company["shares_outstanding_m"], NUM_FMT),
        ("Current Share Price (AUD)", company["share_price_aud"], AUD_FMT),
        ("Market Cap (AUD m)", company["market_cap_aud_m"], NUM_FMT),
        ("AUD/USD FX Rate", company["aud_usd_fx"], '0.000'),
    ]
    for label, value, fmt in snapshot_rows:
        ws.cell(row=row, column=1, value=label).font = BOLD
        c = ws.cell(row=row, column=2, value=value)
        c.font = BLUE
        if fmt:
            c.number_format = fmt
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="VALUATION SUMMARY (AUD per share)").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="Method").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Implied Price (AUD)").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=3, value="Type").font = HEADER_FONT
    ws.cell(row=row, column=3).fill = HEADER_FILL
    row += 1
    val_start = row
    val_rows = [
        ("Current Share Price", dcf["current_share_price_aud"], "Market", BLUE),
        ("DCF Implied Value", dcf["implied_share_price_aud"], "Intrinsic", GREEN),
        ("Comps (EV/EBITDA avg) Implied Value", comps["implied_price_ev_avg_aud"], "Relative", GREEN),
        ("Comps (P/E avg) Implied Value", comps["implied_price_pe_avg_aud"], "Relative", GREEN),
    ]
    for label, value, vtype, font in val_rows:
        ws.cell(row=row, column=1, value=label).font = BOLD
        c = ws.cell(row=row, column=2, value=value)
        c.number_format = AUD_FMT
        c.font = font
        ws.cell(row=row, column=3, value=vtype)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Implied Upside / (Downside) vs Current - DCF").font = BOLD
    dcf_row = val_start + 1
    cell = ws.cell(row=row, column=2, value=f"=(B{dcf_row}-B{val_start})/B{val_start}")
    cell.number_format = PCT_FMT
    cell.font = BLACK
    row += 2

    ws.cell(row=row, column=1, value="KEY DRIVERS (DCF)").font = SECTION_FONT
    row += 1
    driver_rows = [
        ("WACC", dcf["wacc"], PCT_FMT, BLUE),
        ("Terminal Growth Rate", dcf["terminal_growth_rate"], PCT_FMT, BLUE),
        ("Enterprise Value (USDm)", dcf["enterprise_value"], NUM_FMT, GREEN),
        ("Net Debt (USDm)", dcf["net_debt"], NUM_FMT, GREEN),
        ("Equity Value (USDm)", dcf["equity_value"], NUM_FMT, GREEN),
    ]
    for label, value, fmt, font in driver_rows:
        ws.cell(row=row, column=1, value=label).font = BOLD
        c = ws.cell(row=row, column=2, value=value)
        c.number_format = fmt
        c.font = font
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="COMPS SUMMARY").font = SECTION_FONT
    row += 1
    comps_rows = [
        ("BHP Current EV/EBITDA", comps["bhp_current_ev_ebitda"], MULT_FMT),
        ("Peer Average EV/EBITDA", comps["peer_avg_ev_ebitda"], MULT_FMT),
        ("BHP Current P/E", comps["bhp_current_pe"], MULT_FMT),
        ("Peer Average P/E", comps["peer_avg_pe"], MULT_FMT),
    ]
    for label, value, fmt in comps_rows:
        ws.cell(row=row, column=1, value=label).font = BOLD
        c = ws.cell(row=row, column=2, value=value)
        c.number_format = fmt
        c.font = GREEN
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="RECOMMENDATION").font = SECTION_FONT
    row += 1
    rec, rationale, _ = derive_recommendation(dcf, comps)
    rec_cell = ws.cell(row=row, column=1, value=rec)
    rec_cell.font = Font(bold=True, size=14, color="FFFFFF")
    fill_color = {"BUY": "2E7D32", "HOLD": "F9A825", "SELL": "C62828"}[rec]
    rec_cell.fill = PatternFill("solid", start_color=fill_color)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    rat_cell = ws.cell(row=row, column=1, value=rationale)
    rat_cell.font = Font(italic=True, size=9)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 45
    rat_cell.alignment = Alignment(wrap_text=True, vertical="top")

    return ws


def build_three_statement_sheet(wb, model):
    ws = wb.create_sheet("3-Statement Model")
    set_col_widths(ws, [26] + [12] * 7)
    write_title(ws, "3-Statement Financial Model", "USD millions | FY24A-FY25A actuals, FY26E-FY30E forecast")

    row = 4
    ws.cell(row=row, column=1, value="INCOME STATEMENT").font = SECTION_FONT
    row += 1
    df = model["income_statement"].copy()
    df.index.name = "(USDm)"
    row = write_table(ws, row, df)

    ws.cell(row=row, column=1, value="BALANCE SHEET").font = SECTION_FONT
    row += 1
    df = model["balance_sheet"].copy()
    df.index.name = "(USDm)"
    row = write_table(ws, row, df)

    ws.cell(row=row, column=1, value="CASH FLOW STATEMENT").font = SECTION_FONT
    row += 1
    df = model["cash_flow"].copy()
    df.index.name = "(USDm)"
    row = write_table(ws, row, df,
                       source_text="Source: BHP FY2025 Annual Report / SEC Form 6-K (actuals, Aug 2025); "
                                    "forecasts driven by assumptions on the 'Assumptions' tab.")
    return ws


def build_assumptions_sheet(wb, model):
    ws = wb.create_sheet("Assumptions")
    set_col_widths(ws, [30, 12, 12, 12, 12, 12])
    write_title(ws, "Forecast Assumptions", "All blue cells are hardcoded inputs - change these to flex the model", last_col=6)

    a = model["assumptions"]
    row = 4
    ws.cell(row=row, column=1, value="DRIVER").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    for j, fy in enumerate(tsm.FORECAST_YEARS, start=2):
        c = ws.cell(row=row, column=j, value=fy)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    row += 1

    ws.cell(row=row, column=1, value="Revenue Growth %").font = BOLD
    for j, v in enumerate(a["revenue_growth"], start=2):
        c = ws.cell(row=row, column=j, value=v)
        c.number_format = PCT_FMT
        c.font = BLUE
    row += 1

    ws.cell(row=row, column=1, value="EBITDA Margin %").font = BOLD
    for j, v in enumerate(a["ebitda_margin"], start=2):
        c = ws.cell(row=row, column=j, value=v)
        c.number_format = PCT_FMT
        c.font = BLUE
    row += 2

    single_assumptions = [
        ("D&A as % of Revenue", a["da_pct_revenue"], PCT_FMT),
        ("Capex as % of Revenue", a["capex_pct_revenue"], PCT_FMT),
        ("Effective Tax Rate", a["tax_rate"], PCT_FMT),
        ("Net Debt Target (USDm)", a["net_debt_target"], NUM_FMT),
        ("Shares Outstanding (m)", a["shares_outstanding_m"], NUM_FMT),
        ("Dividend Payout Ratio", 0.55, PCT_FMT),
    ]
    for label, value, fmt in single_assumptions:
        ws.cell(row=row, column=1, value=label).font = BOLD
        c = ws.cell(row=row, column=2, value=value)
        c.number_format = fmt
        c.font = BLUE
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="NOTES & SOURCING").font = SECTION_FONT
    row += 1
    notes = [
        "Revenue growth tapers to ~2.5% (mature large-cap miner; no major new mines beyond Jansen ramp-up).",
        "EBITDA margin normalises from FY25's 51% toward BHP's ~20-year average of ~50% as commodity prices ease.",
        "Capex elevated at 17% of revenue, reflecting BHP's stated FY26-27 capex guidance (~US$11bn/yr) for the Jansen potash project.",
        "Net debt held within BHP's stated target range of US$10-20bn.",
        "Cash flow forecast is simplified: ignores working capital swings, JV distributions and divestment proceeds.",
        "Historical actuals sourced from BHP FY2025 Annual Report and SEC Form 6-K (August 2025).",
    ]
    for n in notes:
        c = ws.cell(row=row, column=1, value="- " + n)
        c.font = Font(size=9, italic=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

    return ws


def build_dcf_sheet(wb, model, dcf):
    ws = wb.create_sheet("DCF")
    set_col_widths(ws, [28] + [12] * 6)
    write_title(ws, "Discounted Cash Flow (DCF) Valuation",
                 "USD millions unless stated | 5-year explicit forecast + Gordon Growth terminal value", last_col=7)

    row = 4
    ws.cell(row=row, column=1, value="WACC BUILD-UP").font = SECTION_FONT
    row += 1
    wacc_b = dcf["wacc_breakdown"]
    wacc_start = row
    wacc_rows = [
        ("Risk-Free Rate", val.WACC_ASSUMPTIONS["risk_free_rate"], PCT_FMT, True),
        ("Beta", val.WACC_ASSUMPTIONS["beta"], '0.00', True),
        ("Equity Risk Premium", val.WACC_ASSUMPTIONS["equity_risk_premium"], PCT_FMT, True),
        ("Cost of Equity (CAPM)", None, PCT_FMT, False),
        ("Cost of Debt (pre-tax)", val.WACC_ASSUMPTIONS["cost_of_debt"], PCT_FMT, True),
        ("Tax Rate", val.WACC_ASSUMPTIONS["tax_rate"], PCT_FMT, True),
        ("Cost of Debt (after-tax)", None, PCT_FMT, False),
        ("Weight of Equity (E/V)", wacc_b["weight_equity"], PCT_FMT, True),
        ("Weight of Debt (D/V)", wacc_b["weight_debt"], PCT_FMT, True),
        ("WACC", None, PCT_FMT, False),
    ]
    cell_rows = {}
    for label, value, fmt, is_input in wacc_rows:
        cell_rows[label] = row
        ws.cell(row=row, column=1, value=label).font = BOLD
        c = ws.cell(row=row, column=2)
        c.number_format = fmt
        if is_input:
            c.value = value
            c.font = BLUE
        else:
            c.font = BLACK
        row += 1

    # Now fill formula cells with cross-references
    rfr_r = cell_rows["Risk-Free Rate"]
    beta_r = cell_rows["Beta"]
    erp_r = cell_rows["Equity Risk Premium"]
    coe_r = cell_rows["Cost of Equity (CAPM)"]
    cod_r = cell_rows["Cost of Debt (pre-tax)"]
    tax_r = cell_rows["Tax Rate"]
    coda_r = cell_rows["Cost of Debt (after-tax)"]
    we_r = cell_rows["Weight of Equity (E/V)"]
    wd_r = cell_rows["Weight of Debt (D/V)"]
    wacc_r = cell_rows["WACC"]

    ws.cell(row=coe_r, column=2, value=f"=B{rfr_r}+B{beta_r}*B{erp_r}")
    ws.cell(row=coda_r, column=2, value=f"=B{cod_r}*(1-B{tax_r})")
    wacc_cell = ws.cell(row=wacc_r, column=2, value=f"=B{we_r}*B{coe_r}+B{wd_r}*B{coda_r}")
    wacc_cell.font = Font(bold=True)
    ws.cell(row=wacc_r, column=1).font = Font(bold=True)
    wacc_cell_row = wacc_r

    row += 1
    ws.cell(row=row, column=1, value="FREE CASH FLOW FORECAST & DISCOUNTING").font = SECTION_FONT
    row += 1
    forecast_years = [c for c in model["cash_flow"].columns if c.endswith("E")]
    headers = ["(USDm)"] + forecast_years
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center" if j > 1 else "left")
    row += 1

    ws.cell(row=row, column=1, value="Free Cash Flow").font = BOLD
    for j, fy in enumerate(forecast_years, start=2):
        c = ws.cell(row=row, column=j, value=float(model["cash_flow"].loc["Free Cash Flow", fy]))
        c.number_format = NUM_FMT
        c.font = GREEN
    fcf_row = row
    row += 1

    ws.cell(row=row, column=1, value="Discount Period (n)").font = BOLD
    for j in range(2, 2 + len(forecast_years)):
        c = ws.cell(row=row, column=j, value=j - 1)
        c.number_format = '0'
        c.font = BLACK
    period_row = row
    row += 1

    ws.cell(row=row, column=1, value="Discount Factor").font = BOLD
    for j in range(2, 2 + len(forecast_years)):
        col = get_column_letter(j)
        c = ws.cell(row=row, column=j, value=f"=1/(1+$B${wacc_cell_row})^{col}{period_row}")
        c.number_format = '0.000'
        c.font = BLACK
    df_row = row
    row += 1

    ws.cell(row=row, column=1, value="PV of Free Cash Flow").font = BOLD
    for j in range(2, 2 + len(forecast_years)):
        col = get_column_letter(j)
        c = ws.cell(row=row, column=j, value=f"={col}{fcf_row}*{col}{df_row}")
        c.number_format = NUM_FMT
        c.font = BLACK
    pv_row = row
    row += 2

    ws.cell(row=row, column=1, value="Sum of PV of FCFs").font = BOLD
    last_col = get_column_letter(1 + len(forecast_years))
    sum_pv_cell = ws.cell(row=row, column=2, value=f"=SUM(B{pv_row}:{last_col}{pv_row})")
    sum_pv_cell.number_format = NUM_FMT
    sum_pv_cell.font = BLACK
    sum_pv_row = row
    row += 1

    ws.cell(row=row, column=1, value="Terminal Growth Rate (g)").font = BOLD
    g_cell = ws.cell(row=row, column=2, value=dcf["terminal_growth_rate"])
    g_cell.number_format = PCT_FMT
    g_cell.font = BLUE
    g_row = row
    row += 1

    ws.cell(row=row, column=1, value="Terminal Value (undiscounted)").font = BOLD
    last_fcf_col = get_column_letter(1 + len(forecast_years))
    tv_cell = ws.cell(row=row, column=2,
                       value=f"={last_fcf_col}{fcf_row}*(1+B{g_row})/(B{wacc_cell_row}-B{g_row})")
    tv_cell.number_format = NUM_FMT
    tv_cell.font = BLACK
    tv_row = row
    row += 1

    ws.cell(row=row, column=1, value="PV of Terminal Value").font = BOLD
    last_df_col = get_column_letter(1 + len(forecast_years))
    pv_tv_cell = ws.cell(row=row, column=2, value=f"=B{tv_row}*{last_df_col}{df_row}")
    pv_tv_cell.number_format = NUM_FMT
    pv_tv_cell.font = BLACK
    pv_tv_row = row
    row += 2

    ws.cell(row=row, column=1, value="Enterprise Value").font = Font(bold=True, size=12)
    ev_cell = ws.cell(row=row, column=2, value=f"=B{sum_pv_row}+B{pv_tv_row}")
    ev_cell.number_format = NUM_FMT
    ev_cell.font = Font(bold=True)
    ev_row = row
    row += 1

    ws.cell(row=row, column=1, value="(-) Net Debt").font = BOLD
    nd_cell = ws.cell(row=row, column=2, value=dcf["net_debt"])
    nd_cell.number_format = '(#,##0);#,##0'
    nd_cell.font = GREEN
    nd_row = row
    row += 1

    ws.cell(row=row, column=1, value="Equity Value").font = Font(bold=True, size=12)
    eqv_cell = ws.cell(row=row, column=2, value=f"=B{ev_row}-B{nd_row}")
    eqv_cell.number_format = NUM_FMT
    eqv_cell.font = Font(bold=True)
    eqv_row = row
    row += 1

    ws.cell(row=row, column=1, value="Shares Outstanding (m)").font = BOLD
    sh_cell = ws.cell(row=row, column=2, value=5382.6)
    sh_cell.number_format = NUM_FMT
    sh_cell.font = BLUE
    sh_row = row
    row += 1

    ws.cell(row=row, column=1, value="Implied Value per Share (USD)").font = Font(bold=True, size=12)
    ps_usd_cell = ws.cell(row=row, column=2, value=f"=B{eqv_row}/B{sh_row}")
    ps_usd_cell.number_format = '$#,##0.00'
    ps_usd_cell.font = Font(bold=True)
    ps_usd_row = row
    row += 1

    ws.cell(row=row, column=1, value="AUD/USD FX Rate").font = BOLD
    fx_cell = ws.cell(row=row, column=2, value=0.655)
    fx_cell.number_format = '0.000'
    fx_cell.font = BLUE
    fx_row = row
    row += 1

    ws.cell(row=row, column=1, value="Implied Value per Share (AUD)").font = Font(bold=True, size=13, color="1F4E78")
    ps_aud_cell = ws.cell(row=row, column=2, value=f"=B{ps_usd_row}/B{fx_row}")
    ps_aud_cell.number_format = AUD_FMT
    ps_aud_cell.font = Font(bold=True, size=13, color="1F4E78")
    ps_aud_cell.fill = PatternFill("solid", start_color="D9E1F2")
    implied_aud_row = row
    row += 1

    ws.cell(row=row, column=1, value="Current Share Price (AUD)").font = BOLD
    cur_cell = ws.cell(row=row, column=2, value=dcf["current_share_price_aud"])
    cur_cell.number_format = AUD_FMT
    cur_cell.font = BLUE
    cur_row = row
    row += 1

    ws.cell(row=row, column=1, value="Implied Upside / (Downside)").font = Font(bold=True, size=12)
    up_cell = ws.cell(row=row, column=2, value=f"=B{implied_aud_row}/B{cur_row}-1")
    up_cell.number_format = PCT_FMT
    up_cell.font = Font(bold=True, size=12)

    return ws


def build_comps_sheet(wb, comps):
    ws = wb.create_sheet("Comps")
    set_col_widths(ws, [24, 16, 16, 16])
    write_title(ws, "Comparable Company Analysis",
                 "Peer multiples applied to BHP's own EBITDA / EPS to derive implied valuation", last_col=4)

    row = 4
    ws.cell(row=row, column=1, value="PEER GROUP").font = SECTION_FONT
    row += 1
    df = comps["peers"].set_index("Company").drop(columns=["Ticker"])
    headers = ["Company"] + list(df.columns)
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1
    for idx, r in df.iterrows():
        ws.cell(row=row, column=1, value=idx).font = BOLD
        for j, col in enumerate(df.columns, start=2):
            c = ws.cell(row=row, column=j, value=float(r[col]))
            c.number_format = MULT_FMT
            c.font = BLUE
        row += 1
    peer_start = row - len(df)
    peer_end = row - 1

    row += 1
    ws.cell(row=row, column=1, value="Peer Average").font = BOLD
    avg_ev_cell = ws.cell(row=row, column=2, value=f"=AVERAGE(B{peer_start}:B{peer_end})")
    avg_ev_cell.number_format = MULT_FMT
    avg_pe_cell = ws.cell(row=row, column=3, value=f"=AVERAGE(C{peer_start}:C{peer_end})")
    avg_pe_cell.number_format = MULT_FMT
    avg_ev_row = row
    avg_pe_row = row
    row += 1

    ws.cell(row=row, column=1, value="Peer Median").font = BOLD
    med_ev_cell = ws.cell(row=row, column=2, value=f"=MEDIAN(B{peer_start}:B{peer_end})")
    med_ev_cell.number_format = MULT_FMT
    med_pe_cell = ws.cell(row=row, column=3, value=f"=MEDIAN(C{peer_start}:C{peer_end})")
    med_pe_cell.number_format = MULT_FMT
    med_ev_row = row
    med_pe_row = row
    row += 2

    ws.cell(row=row, column=1, value="BHP FUNDAMENTALS").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="BHP EBITDA FY25 (USDm)").font = BOLD
    ebitda_cell = ws.cell(row=row, column=2, value=float(comps["bhp_ebitda_fy25"]))
    ebitda_cell.number_format = NUM_FMT
    ebitda_cell.font = GREEN
    ebitda_row = row
    row += 1

    ws.cell(row=row, column=1, value="BHP EPS FY25 (USD)").font = BOLD
    eps_cell = ws.cell(row=row, column=2, value=float(comps["bhp_eps_usd"]))
    eps_cell.number_format = '$#,##0.00'
    eps_cell.font = GREEN
    eps_row = row
    row += 1

    ws.cell(row=row, column=1, value="Net Debt (USDm)").font = BOLD
    nd_cell = ws.cell(row=row, column=2, value=12924)
    nd_cell.number_format = NUM_FMT
    nd_cell.font = BLUE
    nd_row = row
    row += 1

    ws.cell(row=row, column=1, value="Shares Outstanding (m)").font = BOLD
    sh_cell = ws.cell(row=row, column=2, value=5382.6)
    sh_cell.number_format = NUM_FMT
    sh_cell.font = BLUE
    sh_row = row
    row += 1

    ws.cell(row=row, column=1, value="AUD/USD FX Rate").font = BOLD
    fx_cell = ws.cell(row=row, column=2, value=0.655)
    fx_cell.number_format = '0.000'
    fx_cell.font = BLUE
    fx_row = row
    row += 2

    ws.cell(row=row, column=1, value="IMPLIED VALUATION").font = SECTION_FONT
    row += 1
    headers = ["Method", "Implied Equity Value (USDm)", "Implied Price (USD)", "Implied Price (AUD)"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    row += 1

    ws.cell(row=row, column=1, value="EV/EBITDA (Peer Avg)").font = BOLD
    eq_cell = ws.cell(row=row, column=2, value=f"=B{avg_ev_row}*B{ebitda_row}-B{nd_row}")
    eq_cell.number_format = NUM_FMT
    px_usd = ws.cell(row=row, column=3, value=f"=B{row}/B{sh_row}")
    px_usd.number_format = '$#,##0.00'
    px_aud = ws.cell(row=row, column=4, value=f"=C{row}/B{fx_row}")
    px_aud.number_format = AUD_FMT
    row += 1

    ws.cell(row=row, column=1, value="EV/EBITDA (Peer Median)").font = BOLD
    eq_cell = ws.cell(row=row, column=2, value=f"=B{med_ev_row}*B{ebitda_row}-B{nd_row}")
    eq_cell.number_format = NUM_FMT
    px_usd = ws.cell(row=row, column=3, value=f"=B{row}/B{sh_row}")
    px_usd.number_format = '$#,##0.00'
    px_aud = ws.cell(row=row, column=4, value=f"=C{row}/B{fx_row}")
    px_aud.number_format = AUD_FMT
    row += 1

    ws.cell(row=row, column=1, value="P/E (Peer Avg)").font = BOLD
    ws.cell(row=row, column=2, value="n/a")
    px_usd = ws.cell(row=row, column=3, value=f"=B{avg_pe_row}*B{eps_row}")
    px_usd.number_format = '$#,##0.00'
    px_aud = ws.cell(row=row, column=4, value=f"=C{row}/B{fx_row}")
    px_aud.number_format = AUD_FMT
    row += 1

    ws.cell(row=row, column=1, value="P/E (Peer Median)").font = BOLD
    ws.cell(row=row, column=2, value="n/a")
    px_usd = ws.cell(row=row, column=3, value=f"=B{med_pe_row}*B{eps_row}")
    px_usd.number_format = '$#,##0.00'
    px_aud = ws.cell(row=row, column=4, value=f"=C{row}/B{fx_row}")
    px_aud.number_format = AUD_FMT
    row += 2

    ws.cell(row=row, column=1, value="BHP Current EV/EBITDA").font = BOLD
    c = ws.cell(row=row, column=2, value=float(comps["bhp_current_ev_ebitda"]))
    c.number_format = MULT_FMT
    c.font = GREEN
    row += 1
    ws.cell(row=row, column=1, value="BHP Current P/E").font = BOLD
    c = ws.cell(row=row, column=2, value=float(comps["bhp_current_pe"]))
    c.number_format = MULT_FMT
    c.font = GREEN
    row += 2

    note = ws.cell(row=row, column=1,
                    value="Peer multiples are illustrative analyst-style estimates (mid-2026) for major diversified "
                          "miners; see README for methodology and sourcing notes.")
    note.font = Font(italic=True, size=8, color="808080")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    return ws


def apply_default_font(wb):
    for ws in wb.worksheets:
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                existing = cell.font
                if existing.name != "Arial":
                    cell.font = Font(
                        name="Arial",
                        bold=existing.bold,
                        italic=existing.italic,
                        color=existing.color,
                        size=existing.size or 10,
                    )


def main():
    snapshot = data_fetch.get_financials()
    model = tsm.build_model(snapshot)
    dcf = val.run_dcf(model, snapshot)
    comps = val.run_comps(model, snapshot)

    wb = Workbook()
    wb.remove(wb.active)

    build_dashboard(wb, snapshot, model, dcf, comps)
    build_three_statement_sheet(wb, model)
    build_assumptions_sheet(wb, model)
    build_dcf_sheet(wb, model, dcf)
    build_comps_sheet(wb, comps)

    apply_default_font(wb)

    out_dir = os.path.join(os.path.dirname(__file__), "..", "output")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "BHP_Equity_Research_Model.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")
    return out_path, snapshot, model, dcf, comps


if __name__ == "__main__":
    main()
